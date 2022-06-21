# -- coding: utf-8 --
# @Time : 2022/6/21 11:20
# @Author : siyu.yang
from openpyxl import load_workbook

class HandleExcel:
    def __init__(self,file_name,sheet_name):
        self.workbook_object = load_workbook(filename=file_name)
        self.sheet_object = self.workbook_object[sheet_name]

    def get_excel_test_case(self):
        cookie_list= []
        datas = list(self.sheet_object.iter_rows(values_only=True))
        case_title = datas[0]
        case_datas = datas[1:]
        for case in case_datas:
            result = dict(zip(case_title,case))
            cookie_list.append(result)
        self.close_file()
        return cookie_list

    def close_file(self):
        self.workbook_object.close()

if __name__ == '__main__':
    import os
    excel_path = os.path.join(os.path.abspath(__file__), "../data/zentao_login_cookies.xlsx")
    cl = HandleExcel(file_name="C:\\Users\\kcadmin\\Desktop\\ui_base_projec\\homework\\homework_20220619\\data\\zentao_login_cookies.xlsx",
                     sheet_name='zentao_cookies')
    print(cl.get_excel_test_case())