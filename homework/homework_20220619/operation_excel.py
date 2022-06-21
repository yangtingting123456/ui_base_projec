# -- coding: utf-8 --
# @Time : 2022/6/21 11:01
import openpyxl
class OperationExcel(object):
    """处理Excel数据,并且写⼊
    """
    def __init__(self, sheet):
        """
        根据嵌套字典获取Excel的表头
        :param sheet: 定义⼀个sheet的名称
        """
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.create_sheet(index=0,title=sheet)

    def get_title(self, data):
        """
                :param data: 传⼊json数据的嵌套字典
                :return: 返回所有字典的标题（keys)
                """
        title_list = []
        for dic in data:
            for key in dic:
                if key not in title_list:
                    title_list.append(key)
        return title_list

    def handle_data(self, title_list, data, filename):
        """
        :param title_list:
        :param data: 调⽤get_title返回的标题列表
        :param data: 传⼊json数据的嵌套字典
        :return: 返回所有字典的标题（keys)
        """
        # 1. 把列表的标题插⼊Excel的第⼀⾏
        first_row = 1
        for header in title_list:
            col = title_list.index(header)
            self.ws.cell(first_row, col + 1, header)
        # 2. 把所有的值,根据标题去筛选,插⼊excel
        row = 2
        for player in data:
            for _key, _value in player.items():
                col = title_list.index(_key)
                self.ws.cell(row, col + 1, _value)
            row += 1  # enter the next row
        self.wb.save(filename)
        self.wb.close()




if __name__ == '__main__':
    import os
    excel_path = os.path.join(os.path.abspath(__file__), "../data/zentao_login_cookies.xlsx")
    filename = "zentao_cookies.xlsx"
    sheet = "zentao_cookies"
    data = [{"name": "zentaosid", "value": "l97vmm6ovptiq8c47s6unb2e31", "Domain": "47.107.178.45", "path": "/"},
                {"name": "theme", "value": "default", "Domain": "47.107.178.45", "path": "/zentao/www/", },
                {"name": "device", "value": "desktop", "Domain": "47.107.178.45", "path": "/zentao/www/", },
                {"name": "lang", "value": "zh-cn", "Domain": "47.107.178.45", "path": "/zentao/www/"}]
    A = OperationExcel(sheet)
    title_list = A.get_title(data)
    # 清洗数据，写⼊Excel
    A.handle_data(title_list=title_list,
                  data=data,
                  filename=excel_path)
